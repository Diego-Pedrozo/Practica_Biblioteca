from rest_framework.viewsets import ModelViewSet
from apps.material_bibliografico.models.solicitud import SolicitudModel, LibroModel
from apps.material_bibliografico.serializers.solicitud import SolicitudSerializer, SolicitudCreateSerializer, SolicitudStatusUpdateSerializer, LibroSerializer

from django.contrib.auth.models import User
from apps.user.models.information import UserInformationModel
from rest_framework.permissions import IsAuthenticated
from rest_framework import status
from rest_framework.response import Response
from rest_framework.response import Response
from rest_framework.decorators import action
from datetime import date, datetime
# from django.db.models import Q
import openpyxl
from openpyxl.styles import Font
from django.http import HttpResponse
from apps.user.choices import UserFacultad, UserPrograma, NivelRevision
from apps.material_bibliografico.models.notificacion import NotificacionModel
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
import pandas as pd
from ..paginators import CustomPaginator, CustomPaginatorFiveItems
from apps.material_bibliografico.models.fechas_limite import FechasLimiteModel

class SolicitudPublicViewSet(ModelViewSet):
    model = SolicitudModel
    serializer_class = SolicitudSerializer
    queryset = SolicitudModel.objects.all()    
    http_method_names = ['get', 'post']


    def list(self, request, *args, **kwargs):
        print(self.action)

        paginator = CustomPaginatorFiveItems()
        
        email = request.query_params.get('email')
        if not email:
            return Response({'mensaje': 'El email es requerido'}, status=status.HTTP_400_BAD_REQUEST)
        
        qs = self.get_queryset().filter(email_solicitante=email).order_by('-id')
        result_page = paginator.paginate_queryset(qs, request)
        serializer = self.get_serializer(result_page, many=True) 
    
        if qs.exists():
            return paginator.get_paginated_response(serializer.data)
        else:
            return Response({'mensaje': 'No se encontraron solicitudes para ese email'}, status=status.HTTP_404_NOT_FOUND)

    def create(self, request, *args, **kwargs):
        print(self.action)
        try:
            fecha_actual = date.today()
            fechas_limite = FechasLimiteModel.objects.first()

            if not fechas_limite:
                return Response({'mensaje': 'No hay fechas límite definidas para aceptar solicitudes'}, status=status.HTTP_400_BAD_REQUEST)
            
            if fecha_actual < fechas_limite.fecha_inicio or fecha_actual > fechas_limite.fecha_fin:
                return Response({'mensaje': 'Las solicitudes no están permitidas en esta fecha'}, status=status.HTTP_400_BAD_REQUEST)

            data= request.data
            data_libro = data.get('libro', {})
            data_solicitud = data.get('solicitud', {})

            serializer_libro = LibroSerializer(data=data_libro)
            if serializer_libro.is_valid():
                libro = serializer_libro.save()

            data_solicitud['libro'] = libro.id
            serializer_solicitud = SolicitudCreateSerializer(data=data_solicitud)

            if serializer_solicitud.is_valid():
                solicitud = serializer_solicitud.save()
                return Response({'mensaje': 'Solicitud creada'}, status=status.HTTP_201_CREATED)
            
            else:
                libro.delete()
                return Response({'mensaje': 'Error de solicitud', 'info': serializer_solicitud.errors}, status=status.HTTP_400_BAD_REQUEST)
        
        except Exception as e:
            libro.delete()
            return Response({'mensaje': 'Error al procesar la petición', 'info': {str(e)}}, status=status.HTTP_400_BAD_REQUEST)
        
    @action(detail=False, methods=['post'], url_path='cargar_varias')
    def variasSolicitudes(self, request, *args, **kwargs):
        print("variasSolicitudes()")
        try:
            fecha_actual = date.today()
            fechas_limite = FechasLimiteModel.objects.first()

            if not fechas_limite:
                return Response({'mensaje': 'No hay fechas límite definidas para aceptar solicitudes'}, status=status.HTTP_400_BAD_REQUEST)
            
            if fecha_actual < fechas_limite.fecha_inicio or fecha_actual > fechas_limite.fecha_fin:
                return Response({'mensaje': 'Las solicitudes no están permitidas en esta fecha'}, status=status.HTTP_400_BAD_REQUEST)
            
            file = request.FILES['file']
            facultad = request.data.get('facultad')
            programa_academico = request.data.get('programa_academico')
            solicitante = request.data.get('solicitante')
            emailSolicitante = request.data.get('email_solicitante')

            file_name = default_storage.save(file.name, ContentFile(file.read()))
            file_path = default_storage.path(file_name)

            df = pd.read_excel(file_path)

            for index, row in df.iterrows():
                data_libro = {
                    "titulo": row['Título'],
                    "autor": row['Autor'],
                    "editorial": row['Editorial'],
                    "edicion": row['Edición'],
                    "ejemplares": row['Ejemplares'],
                    "fecha_publicacion": row['Año publicación'],
                    "idioma": row['Idioma (Ingles o Español)']
                }
                serializer_libro = LibroSerializer(data=data_libro)
                if serializer_libro.is_valid():
                    libro = serializer_libro.save()
                else:
                    return Response({'mensaje': 'Error en los datos del libro', 'info': serializer_libro.errors}, status=status.HTTP_400_BAD_REQUEST)

                data_solicitud = {
                    "libro": libro.id,
                    "facultad": facultad,
                    "programa_academico": programa_academico,
                    "solicitante": solicitante,
                    "email_solicitante": emailSolicitante
                }
                serializer_solicitud = SolicitudCreateSerializer(data=data_solicitud)
                if serializer_solicitud.is_valid():
                    serializer_solicitud.save()
                else:
                    libro.delete()
                    return Response({'mensaje': 'Error en los datos de la solicitud', 'info': serializer_solicitud.errors}, status=status.HTTP_400_BAD_REQUEST)

            return Response({'mensaje': 'Solicitudes creadas correctamente'}, status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response({'mensaje': 'Error al procesar la petición', 'info': str(e)}, status=status.HTTP_400_BAD_REQUEST)

class SolicitudViewSet(ModelViewSet):
    model = SolicitudModel
    serializer_class = SolicitudSerializer
    queryset = SolicitudModel.objects.all()    
    permission_classes = [IsAuthenticated]
    http_method_names = ['get', 'patch', 'delete', 'post']

    def list(self, request, *args, **kwargs):
        print(self.action)
        user = self.request.user
        user_information = UserInformationModel.objects.get(user=user)

        queryset = self.filter_queryset(self.get_queryset())

        if user_information.user_type in ['2']:
            print('Director plan de estudios')
            #queryset = self.filter_queryset(self.get_queryset().filter(nivel_revision=1))
            queryset = queryset.filter(nivel_revision=1, solicitante='Estudiante')
        elif user_information.user_type in ['3']:
            print('Director de departamento')
            #queryset = self.filter_queryset(self.get_queryset().filter(nivel_revision=1))
            queryset = queryset.filter(nivel_revision=1, solicitante='Docente')
        elif user_information.user_type in ['4']:
            print('Decano')
            #queryset = self.filter_queryset(self.get_queryset().filter(nivel_revision=2))
            queryset = queryset.filter(nivel_revision=2)
        elif user_information.user_type in ['5']:
            print('Biblioteca')
            #queryset = self.filter_queryset(self.get_queryset().filter(nivel_revision=3))
            queryset = queryset.filter(nivel_revision=3)
        elif user_information.user_type in ['6']:
            print('Vicerrector')
            #queryset = self.filter_queryset(self.get_queryset().filter(nivel_revision=4))
            queryset = queryset.filter(nivel_revision=4)
        else: return Response({'mensaje': 'No tiene permisos para ver solicitudes'}, status=status.HTTP_403_FORBIDDEN)

        facultad = request.query_params.get('facultad', None)
        programa = request.query_params.get('programa', None)
        estado = request.query_params.get('estado', None)
        nivel_revision = request.query_params.get('nivel_revision', None)

        if facultad:
            queryset = queryset.filter(facultad=facultad)
        if programa:
            queryset = queryset.filter(programa_academico=programa)
        if estado:
            queryset = queryset.filter(estado=estado)
        if nivel_revision:
            queryset = queryset.filter(nivel_revision=nivel_revision)

        serializer_solicitud = self.get_serializer(queryset, many=True)
        return Response(serializer_solicitud.data)
    
    @action(detail=False, methods=['get'], url_path='solicitudes_revisadas')
    def solicitudesRevisadas(self, request, *args, **kwargs):
        print('solicitudesRevisadas()')
        user = self.request.user
        user_information = UserInformationModel.objects.get(user=user)
        paginator = CustomPaginator()

        queryset = self.filter_queryset(self.get_queryset())
        facultad = request.query_params.get('facultad', None)
        programa = request.query_params.get('programa', None)
        estado = request.query_params.get('estado', None)
        nivel_revision = request.query_params.get('nivel_revision', None)
        fecha_inicio = request.query_params.get('fecha_inicio', None)
        fecha_fin = request.query_params.get('fecha_fin', None)

        if user_information.user_type in ['2']:
            print('Director plan de estudios')
            #queryset = self.filter_queryset(self.get_queryset().filter(nivel_revision__in=[2,3,4,5]))
            if nivel_revision == '1':
                queryset = queryset.filter(nivel_revision__in=[1], solicitante='Estudiante').order_by('id')
            elif nivel_revision == '2':
                queryset = queryset.filter(nivel_revision__in=[2,3,4], solicitante='Estudiante').order_by('id')
            elif nivel_revision == '5':
                queryset = queryset.filter(nivel_revision__in=[5], solicitante='Estudiante').order_by('id')
            elif nivel_revision == '6':
                queryset = queryset.filter(nivel_revision__in=[6], solicitante='Estudiante').order_by('id')
        elif user_information.user_type in ['3']:
            print('Director de departamento')
            #queryset = self.filter_queryset(self.get_queryset().filter(nivel_revision=1))
            if nivel_revision == '1':
                queryset = queryset.filter(nivel_revision__in=[1], solicitante='Docente').order_by('id')
            elif nivel_revision == '2':
                queryset = queryset.filter(nivel_revision__in=[2,3,4], solicitante='Docente').order_by('id')
            elif nivel_revision == '5':
                queryset = queryset.filter(nivel_revision__in=[5], solicitante='Docente').order_by('id')
            elif nivel_revision == '6':
                queryset = queryset.filter(nivel_revision__in=[6], solicitante='Docente').order_by('id')
        elif user_information.user_type in ['4']:
            print('Decano')
            #queryset = self.filter_queryset(self.get_queryset().filter(nivel_revision__in=[3,4,5]))
            if nivel_revision == '1':
                queryset = queryset.filter(nivel_revision__in=[2]).order_by('id')
            elif nivel_revision == '2':
                queryset = queryset.filter(nivel_revision__in=[3,4]).order_by('id')
            elif nivel_revision == '5':
                queryset = queryset.filter(nivel_revision__in=[5]).order_by('id')
            elif nivel_revision == '6':
                queryset = queryset.filter(nivel_revision__in=[6]).order_by('id')
        elif user_information.user_type in ['5']:
            print('Biblioteca')
            #queryset = self.filter_queryset(self.get_queryset().filter(nivel_revision__in=[5,6]))
            if nivel_revision == '1':
                queryset = queryset.filter(nivel_revision__in=[3]).order_by('id')
            elif nivel_revision == '2':
                queryset = queryset.filter(nivel_revision__in=[4]).order_by('id')
            elif nivel_revision == '5':
                queryset = queryset.filter(nivel_revision__in=[5]).order_by('id')
            elif nivel_revision == '6':
                queryset = queryset.filter(nivel_revision__in=[6]).order_by('id')
        elif user_information.user_type in ['6']:
            print('Vicerrector')
            #queryset = self.filter_queryset(self.get_queryset().filter(nivel_revision__in=[5,6]))
            if nivel_revision == '1':
                queryset = queryset.filter(nivel_revision__in=[4]).order_by('id')
            elif nivel_revision == '5':
                queryset = queryset.filter(nivel_revision__in=[5]).order_by('id')
            elif nivel_revision == '6':
                queryset = queryset.filter(nivel_revision__in=[6]).order_by('id')
        else: return Response({'mensaje': 'No tiene permisos para ver solicitudes'}, status=status.HTTP_403_FORBIDDEN)

        if facultad:
            queryset = queryset.filter(facultad=facultad)
        if programa:
            queryset = queryset.filter(programa_academico=programa)
        if estado:
            queryset = queryset.filter(estado=estado)
        if fecha_inicio and fecha_fin:
            queryset = queryset.filter(fecha_solicitud__range=[fecha_inicio, fecha_fin])
        elif fecha_inicio:
            queryset = queryset.filter(fecha_solicitud__gte=fecha_inicio)
        elif fecha_fin:
            queryset = queryset.filter(fecha_solicitud__lte=fecha_fin)
        # if nivel_revision in ['5','6']:
        #     queryset = queryset.filter(nivel_revision=nivel_revision)

        result_page = paginator.paginate_queryset(queryset, request)
        serializer_solicitud = self.get_serializer(result_page, many=True)
        # serializer_solicitud = self.get_serializer(queryset, many=True)
        # return Response(serializer_solicitud.data)
        return paginator.get_paginated_response(serializer_solicitud.data)
    
    def partial_update(self, request, *args, **kwargs):
        print(self.action)
        try:
            instance = self.get_object()
            user = self.request.user
            user_information = UserInformationModel.objects.get(user=user)
            data_estado = request.data
            serializer_solicitud = SolicitudStatusUpdateSerializer(data=data_estado)

            if user_information.user_type not in ['5']:
                return Response({'mensaje': 'No tiene permisos para cambiar el estado de la solicitud'}, status=status.HTTP_403_FORBIDDEN)
            
            else:
                if serializer_solicitud.is_valid():
                    instance.estado = data_estado.get('estado')
                    instance.save()
                    
                    if(instance.estado in ['Existente', 'En tramite']):
                        facultad = instance.facultad
                        descripcion = f'El libro {instance.libro.titulo} se encuentra en estado: {instance.estado}.'
                        NotificacionModel.objects.create(
                            descripcion=descripcion,
                            destinario='Decano',
                            facultad=facultad
                        )

                    return Response({'mensaje': 'Estado actualizado'}, status=status.HTTP_200_OK)
                else:
                    return Response({'mensaje': 'Estado no valido'}, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response({'mensaje': {str(e)}}, status=status.HTTP_400_BAD_REQUEST)
        
        
    def destroy(self, request, *args, **kwargs):
        print(self.action)
        instance = self.get_object()

        try:
            libro = instance.libro
        except LibroModel.DoesNotExist:
            libro = None
        
        if libro:
            libro.delete()
        
        instance.delete()

        return Response({'mensaje': 'Solicitud eliminada'}, status=status.HTTP_200_OK)
    
    
    @action(detail=False, methods=['post'], url_path='enviar_solicitudes')
    def solicitudesSeleccionadas(self, request):
        print('solicitudesSeleccionadas()') 
        data = self.request.data

        user = self.request.user
        user_information = UserInformationModel.objects.get(user=user)

        ids_solicitudes = data.get('ids_solicitudes', [])
        if not ids_solicitudes:
            return Response({'mensaje': 'Se requiere al menos un ID de solicitud'}, status=status.HTTP_400_BAD_REQUEST)

        if user_information.user_type in ['2', '3']:
            solicitudes_a_actualizar = SolicitudModel.objects.filter(id__in=ids_solicitudes, nivel_revision='1').values_list('id', flat=True)
            nuevo_nivel_revision = '2'
        elif user_information.user_type in ['4']:
            solicitudes_a_actualizar = SolicitudModel.objects.filter(id__in=ids_solicitudes, nivel_revision='2').values_list('id', flat=True)
            nuevo_nivel_revision = '3'
        elif user_information.user_type in ['5']:
            solicitudes_a_actualizar = SolicitudModel.objects.filter(id__in=ids_solicitudes, nivel_revision='3').values_list('id', flat=True)
            nuevo_nivel_revision = '5'
        elif user_information.user_type in ['6']:
            solicitudes_a_actualizar = SolicitudModel.objects.filter(id__in=ids_solicitudes, nivel_revision='4').values_list('id', flat=True)
            nuevo_nivel_revision = '5'
        else: nuevo_nivel_revision = None
            
        if nuevo_nivel_revision is None:
            return Response({'mensaje': 'Se requiere el nuevo valor de nivel_revision'}, status=status.HTTP_400_BAD_REQUEST)

        solicitudes_a_actualizar = SolicitudModel.objects.filter(id__in=solicitudes_a_actualizar)
        solicitudes_a_actualizar.update(nivel_revision=nuevo_nivel_revision)

        return Response({'mensaje': 'Solicitudes enviadas'}, status=status.HTTP_200_OK)
    
    @action(detail=False, methods=['post'], url_path='rechazar_solicitudes')
    def solicitudesRechazadas(self, request):
        print('solicitudesRechazadas()') 
        data = self.request.data

        user = self.request.user
        user_information = UserInformationModel.objects.get(user=user)

        ids_solicitudes = data.get('ids_solicitudes', [])
        if not ids_solicitudes:
            return Response({'mensaje': 'Se requiere al menos un ID de solicitud'}, status=status.HTTP_400_BAD_REQUEST)
        
        if user_information.user_type not in ['5']:
                return Response({'mensaje': 'No tiene permisos para cambiar el estado de la solicitud'}, status=status.HTTP_403_FORBIDDEN)

        solicitudes_a_actualizar = SolicitudModel.objects.filter(id__in=ids_solicitudes, nivel_revision='3').values_list('id', flat=True)
        nuevo_nivel_revision = '6'
            
        if nuevo_nivel_revision is None:
            return Response({'mensaje': 'Se requiere el nuevo valor de nivel_revision'}, status=status.HTTP_400_BAD_REQUEST)
        
        if not solicitudes_a_actualizar:
            return Response({'mensaje': 'No hay solicitudes para actualizar'}, status=status.HTTP_400_BAD_REQUEST)

        solicitudes_a_actualizar = SolicitudModel.objects.filter(id__in=solicitudes_a_actualizar)
        solicitudes_a_actualizar.update(nivel_revision=nuevo_nivel_revision)

        return Response({'mensaje': 'Solicitudes rechazadas'}, status=status.HTTP_200_OK)
    
    @action(detail=False, methods=['get'], url_path='generar_reporte')
    def generarReporte(self, request):
        print('generarReporte()') 

        user = self.request.user
        user_information = UserInformationModel.objects.get(user=user)

        queryset = self.filter_queryset(self.get_queryset())

        if user_information.user_type in ['2']:
            print('Director plan de estudios')
            queryset_estudiantes = queryset.filter(nivel_revision__in=[1,5,6], solicitante='Estudiante')
        elif user_information.user_type in ['3']:
            print('Director de departamento')
            queryset_docentes = queryset.filter(nivel_revision__in=[1,5,6], solicitante='Docente')
        elif user_information.user_type in ['4']:
            print('Decano')
            #queryset = queryset.filter(nivel_revision=2)
            queryset_estudiantes = queryset.filter(nivel_revision__in=[2,5,6], solicitante='Estudiante')
            queryset_docentes = queryset.filter(nivel_revision__in=[2,5,6], solicitante='Docente')
        elif user_information.user_type in ['5']:
            print('Biblioteca')
            #queryset = queryset.filter(nivel_revision=3)
            queryset_estudiantes = queryset.filter(nivel_revision__in=[3,5,6], solicitante='Estudiante')
            queryset_docentes = queryset.filter(nivel_revision__in=[3,5,6], solicitante='Docente')
        elif user_information.user_type in ['6']:
            print('Vicerrector')
            #queryset = queryset.filter(nivel_revision=4)
            queryset_estudiantes = queryset.filter(nivel_revision__in=[4,5,6], solicitante='Estudiante')
            queryset_docentes = queryset.filter(nivel_revision__in=[4,5,6], solicitante='Docente')
        else: return Response({'mensaje': 'No tiene permisos para ver solicitudes'}, status=status.HTTP_403_FORBIDDEN)

        facultad = request.query_params.get('facultad', None)
        programa = request.query_params.get('programa', None)
        estado = request.query_params.get('estado', None)
        nivel_revision = request.query_params.get('nivel_revision', None)
        fecha_inicio = request.query_params.get('fecha_inicio', None)
        fecha_fin = request.query_params.get('fecha_fin', None)

        if facultad:
            #queryset = queryset.filter(facultad=facultad)
            if user_information.user_type in ['2', '4', '5']:
                queryset_estudiantes = queryset_estudiantes.filter(facultad=facultad)
            if user_information.user_type in ['3', '4', '5']:
                queryset_docentes = queryset_docentes.filter(facultad=facultad)
        if programa:
            #queryset = queryset.filter(programa_academico=programa)
            if user_information.user_type in ['2', '4', '5']:
                queryset_estudiantes = queryset_estudiantes.filter(programa_academico=programa)
            if user_information.user_type in ['3', '4', '5']:
                queryset_docentes = queryset_docentes.filter(programa_academico=programa)
        if estado:
            #queryset = queryset.filter(estado=estado)
            if user_information.user_type in ['2', '4', '5']:
                queryset_estudiantes = queryset_estudiantes.filter(estado=estado)
            if user_information.user_type in ['3', '4', '5']:
                queryset_docentes = queryset_docentes.filter(estado=estado)
        if nivel_revision:
            #queryset = queryset.filter(nivel_revision=nivel_revision)
            if user_information.user_type in ['2', '4', '5']:
                queryset_estudiantes = queryset_estudiantes.filter(nivel_revision=nivel_revision)
            if user_information.user_type in ['3', '4', '5']:
                queryset_docentes = queryset_docentes.filter(nivel_revision=nivel_revision)
        if fecha_inicio and fecha_fin:
            if user_information.user_type in ['2', '4', '5']:
                queryset_estudiantes = queryset_estudiantes.filter(fecha_solicitud__range=[fecha_inicio, fecha_fin])
            if user_information.user_type in ['3', '4', '5']:
                queryset_docentes = queryset_docentes.filter(fecha_solicitud__range=[fecha_inicio, fecha_fin])
        elif fecha_inicio:
            if user_information.user_type in ['2', '4', '5']:
                queryset_estudiantes = queryset_estudiantes.filter(fecha_solicitud__gte=fecha_inicio)
            if user_information.user_type in ['3', '4', '5']:
                queryset_docentes = queryset_docentes.filter(fecha_solicitud__gte=fecha_inicio)
        elif fecha_fin:
            if user_information.user_type in ['2', '4', '5']:
                queryset_estudiantes = queryset_estudiantes.filter(fecha_solicitud__lte=fecha_fin)
            if user_information.user_type in ['3', '4', '5']:
                queryset_docentes = queryset_docentes.filter(fecha_solicitud__lte=fecha_fin)

        # Crear un libro de trabajo y una hoja de cálculo
        workbook = openpyxl.Workbook()
        worksheet_estudiantes  = workbook.active
        worksheet_estudiantes.title = 'Solicitudes Estudiantes'
        worksheet_docentes = workbook.create_sheet(title='Solicitudes Docentes')

        # Definir los encabezados de la hoja de cálculo
        headers = ['ID', 'Solicitante', 'Facultad', 'Programa Académico', 'Estado', 'Nivel de Revisión', 'Fecha solicitud', 'Título', 'Autor', 'Editorial', 'Edición', 'Ejemplares', 'Año publicación', 'Idioma']
        worksheet_estudiantes.append(headers)
        worksheet_docentes.append(headers)

        # Aplicar estilo a los encabezados
        for col in range(1, len(headers) + 1):
            cell_estudiantes = worksheet_estudiantes.cell(row=1, column=col)
            cell_estudiantes.font = Font(bold=True)
            cell_docentes = worksheet_docentes.cell(row=1, column=col)
            cell_docentes.font = Font(bold=True)

         # Agregar datos a la hoja de estudiantes
        if user_information.user_type in ['2', '4', '5']:
            for solicitud in queryset_estudiantes:
                row = [
                    solicitud.id,
                    solicitud.solicitante,
                    UserFacultad(solicitud.facultad).label,
                    UserPrograma(solicitud.programa_academico).label,
                    solicitud.estado,
                    NivelRevision(solicitud.nivel_revision).label,
                    solicitud.fecha_solicitud,
                    solicitud.libro.titulo,
                    solicitud.libro.autor,
                    solicitud.libro.editorial,
                    solicitud.libro.edicion,
                    solicitud.libro.ejemplares,
                    solicitud.libro.fecha_publicacion,
                    solicitud.libro.idioma,
                ]
                worksheet_estudiantes.append(row)

        # Agregar datos a la hoja de docentes
        if user_information.user_type in ['3', '4', '5']:
            for solicitud in queryset_docentes:
                row = [
                    solicitud.id,
                    solicitud.solicitante,
                    UserFacultad(solicitud.facultad).label,
                    UserPrograma(solicitud.programa_academico).label,
                    solicitud.estado,
                    NivelRevision(solicitud.nivel_revision).label,
                    solicitud.fecha_solicitud,
                    solicitud.libro.titulo,
                    solicitud.libro.autor,
                    solicitud.libro.editorial,
                    solicitud.libro.edicion,
                    solicitud.libro.ejemplares,
                    solicitud.libro.fecha_publicacion,
                    solicitud.libro.idioma,
                ]
                worksheet_docentes.append(row)

        # Guardar el archivo Excel en un objeto BytesIO
        from io import BytesIO
        excel_file = BytesIO()
        workbook.save(excel_file)
        excel_file.seek(0)

        # Preparar la respuesta HTTP con el archivo Excel
        response = HttpResponse(excel_file.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=reporte_solicitudes.xlsx'
        return response

